import streamlit as st
import os
import tempfile
import zipfile
from PIL import Image
import numpy as np
from openpyxl import Workbook
from openpyxl.drawing.image import Image as ExcelImage
from io import BytesIO
import pandas as pd
import plotly.graph_objects as go
from pathlib import Path
import time
import shutil
import cv2

# Try to import face_recognition, provide fallback if not available
try:
    import face_recognition
    FACE_RECOGNITION_AVAILABLE = True
except ImportError as e:
    st.error(f"Face recognition library not available: {str(e)}")
    st.info("Please make sure all dependencies are properly installed.")
    FACE_RECOGNITION_AVAILABLE = False

# Page configuration
st.set_page_config(
    page_title="Face Grouping Pro",
    page_icon="👥",
    layout="wide",
    initial_sidebar_state="expanded"
)

# Custom CSS
st.markdown("""
<style>
    .main-header {
        background: linear-gradient(135deg, #4b8bec 0%, #2a5f9e 100%);
        padding: 2rem;
        border-radius: 10px;
        margin-bottom: 2rem;
        color: white;
        text-align: center;
    }
    .stProgress > div > div > div > div {
        background-color: #5cb85c;
    }
    .info-box {
        background-color: #f0f8ff;
        padding: 1rem;
        border-radius: 5px;
        border-left: 5px solid #4b8bec;
        margin-bottom: 1rem;
    }
    .success-box {
        background-color: #dff0d8;
        padding: 1rem;
        border-radius: 5px;
        border-left: 5px solid #5cb85c;
        margin-bottom: 1rem;
    }
    .warning-box {
        background-color: #fcf8e3;
        padding: 1rem;
        border-radius: 5px;
        border-left: 5px solid #f0ad4e;
        margin-bottom: 1rem;
    }
    .footer {
        text-align: center;
        padding: 1rem;
        background: linear-gradient(135deg, #4b8bec 0%, #2a5f9e 100%);
        color: white;
        border-radius: 5px;
        margin-top: 2rem;
    }
</style>
""", unsafe_allow_html=True)

# Header
st.markdown("""
<div class="main-header">
    <h1>👥 Face Grouping Professional</h1>
    <p>Upload images and group similar faces automatically</p>
</div>
""", unsafe_allow_html=True)

# Initialize session state
if 'processed' not in st.session_state:
    st.session_state.processed = False
if 'grouped_faces' not in st.session_state:
    st.session_state.grouped_faces = None
if 'uploaded_files' not in st.session_state:
    st.session_state.uploaded_files = []

# Check if face_recognition is available
if not FACE_RECOGNITION_AVAILABLE:
    st.markdown("""
    <div class="warning-box">
        <h4>⚠️ Face Recognition Library Not Available</h4>
        <p>The face recognition library could not be loaded. This might be due to installation issues.</p>
        <p>Please try one of these solutions:</p>
        <ul>
            <li>Wait a few minutes and refresh the page (dependencies might still be installing)</li>
            <li>Check the app logs for more details</li>
            <li>Contact the app administrator</li>
        </ul>
    </div>
    """, unsafe_allow_html=True)

# ========== Helper Functions ==========
def extract_face_descriptor(image_bytes, image_name):
    """Extract face encoding from image bytes"""
    if not FACE_RECOGNITION_AVAILABLE:
        return []
    
    try:
        # Convert bytes to numpy array
        nparr = np.frombuffer(image_bytes, np.uint8)
        img = cv2.imdecode(nparr, cv2.IMREAD_COLOR)
        
        # Convert BGR to RGB
        img_rgb = cv2.cvtColor(img, cv2.COLOR_BGR2RGB)
        
        # Get face encodings
        face_locations = face_recognition.face_locations(img_rgb)
        encodings = face_recognition.face_encodings(img_rgb, face_locations)
        
        return [(encoding, image_name) for encoding in encodings]
    except Exception as e:
        st.error(f"Error processing {image_name}: {str(e)}")
        return []

def group_similar_embeddings(descriptors, threshold=0.3):
    """Group similar face embeddings"""
    if not descriptors:
        return []
    
    groups = []
    used = [False] * len(descriptors)
    
    for i in range(len(descriptors)):
        if used[i]:
            continue
        group = [descriptors[i][1]]  # Store image name
        used[i] = True
        for j in range(i + 1, len(descriptors)):
            if not used[j]:
                dist = np.linalg.norm(descriptors[i][0] - descriptors[j][0])
                if dist < threshold:
                    group.append(descriptors[j][1])
                    used[j] = True
        if len(group) > 1:  # Only groups with multiple faces
            groups.append(group)
    return groups

def create_excel_with_images(groups, image_data):
    """Create Excel file with grouped faces and embedded images"""
    wb = Workbook()
    ws = wb.active
    ws.title = "Grouped Faces"
    
    # Add headers
    ws['A1'] = "Image"
    ws['B1'] = "Filename"
    ws['C1'] = "Group Number"
    
    row_num = 2
    temp_dir = tempfile.mkdtemp()
    
    try:
        for group_num, group in enumerate(groups, start=1):
            for image_name in group:
                # Save image temporarily
                img_bytes = image_data[image_name]
                temp_path = os.path.join(temp_dir, image_name)
                with open(temp_path, 'wb') as f:
                    f.write(img_bytes)
                
                # Add to Excel
                try:
                    img = ExcelImage(temp_path)
                    img.width = 100
                    img.height = 100
                    ws.add_image(img, f'A{row_num}')
                except:
                    ws[f'A{row_num}'] = "[Image Error]"
                
                ws[f'B{row_num}'] = image_name
                ws[f'C{row_num}'] = f"Group {group_num}"
                row_num += 1
        
        # Save to BytesIO
        excel_bytes = BytesIO()
        wb.save(excel_bytes)
        excel_bytes.seek(0)
        
        return excel_bytes
    
    finally:
        # Cleanup temp directory
        shutil.rmtree(temp_dir, ignore_errors=True)

# ========== Main App Layout ==========
col1, col2 = st.columns([1, 1])

with col1:
    st.markdown("### 📤 Upload Images")
    st.markdown('<div class="info-box">Upload multiple images containing faces. Supported formats: JPG, JPEG, PNG</div>', 
                unsafe_allow_html=True)
    
    uploaded_files = st.file_uploader(
        "Choose images...",
        type=['jpg', 'jpeg', 'png'],
        accept_multiple_files=True,
        key="file_uploader"
    )
    
    if uploaded_files:
        st.session_state.uploaded_files = uploaded_files
        st.success(f"✅ {len(uploaded_files)} images uploaded successfully!")
        
        # Display uploaded images in a grid
        st.markdown("### 👁️ Preview")
        cols = st.columns(4)
        for idx, file in enumerate(uploaded_files[:8]):  # Show first 8 images
            with cols[idx % 4]:
                try:
                    image = Image.open(file)
                    st.image(image, caption=file.name, use_column_width=True)
                except:
                    pass
        
        if len(uploaded_files) > 8:
            st.info(f"... and {len(uploaded_files) - 8} more images")

with col2:
    st.markdown("### ⚙️ Processing Settings")
    
    if FACE_RECOGNITION_AVAILABLE:
        threshold = st.slider(
            "Similarity Threshold",
            min_value=0.1,
            max_value=0.8,
            value=0.3,
            step=0.05,
            help="Lower values = stricter matching, Higher values = looser matching"
        )
        
        use_multiprocessing = st.checkbox("Use multiprocessing (faster)", value=True)
    else:
        st.warning("Face recognition is currently unavailable. Please check the system status.")
    
    st.markdown("### 📊 Statistics")
    
    if st.session_state.processed and st.session_state.grouped_faces:
        total_faces = sum(len(group) for group in st.session_state.grouped_faces)
        total_groups = len(st.session_state.grouped_faces)
        
        # Create metrics
        col2_1, col2_2, col2_3 = st.columns(3)
        col2_1.metric("Total Groups", total_groups)
        col2_2.metric("Total Faces", total_faces)
        col2_3.metric("Avg Faces/Group", f"{(total_faces/total_groups):.1f}")
        
        # Create pie chart
        fig = go.Figure(data=[go.Pie(
            labels=[f"Group {i+1}" for i in range(total_groups)],
            values=[len(group) for group in st.session_state.grouped_faces],
            hole=.3
        )])
        fig.update_layout(height=300)
        st.plotly_chart(fig, use_container_width=True)

# Process button
if st.button("🚀 Start Processing", type="primary", use_container_width=True):
    if not uploaded_files:
        st.error("Please upload at least one image first!")
    elif not FACE_RECOGNITION_AVAILABLE:
        st.error("Face recognition is not available. Cannot process images.")
    else:
        # Progress bar
        progress_bar = st.progress(0)
        status_text = st.empty()
        
        try:
            # Step 1: Extract face encodings
            status_text.text("📸 Extracting faces from images...")
            progress_bar.progress(10)
            
            all_descriptors = []
            image_data = {}
            
            for idx, file in enumerate(uploaded_files):
                # Update progress
                progress = 10 + int((idx / len(uploaded_files)) * 50)
                progress_bar.progress(progress)
                status_text.text(f"Processing {file.name}...")
                
                # Read file bytes
                file_bytes = file.read()
                image_data[file.name] = file_bytes
                
                # Extract face descriptors
                descriptors = extract_face_descriptor(file_bytes, file.name)
                all_descriptors.extend(descriptors)
                
                # Reset file pointer for future reads
                file.seek(0)
            
            if not all_descriptors:
                st.warning("No faces found in the uploaded images!")
            else:
                # Step 2: Group similar faces
                status_text.text("🔄 Grouping similar faces...")
                progress_bar.progress(70)
                
                grouped_faces = group_similar_embeddings(all_descriptors, threshold)
                
                # Step 3: Create Excel file
                status_text.text("📊 Creating Excel report...")
                progress_bar.progress(85)
                
                excel_file = create_excel_with_images(grouped_faces, image_data)
                
                # Step 4: Complete
                progress_bar.progress(100)
                status_text.text("✅ Processing complete!")
                
                # Save to session state
                st.session_state.processed = True
                st.session_state.grouped_faces = grouped_faces
                
                # Success message
                st.markdown(f"""
                <div class="success-box">
                    <h4>✅ Processing Complete!</h4>
                    <p>Found {len(grouped_faces)} groups of similar faces</p>
                </div>
                """, unsafe_allow_html=True)
                
                # Download button
                st.download_button(
                    label="📥 Download Excel Report",
                    data=excel_file,
                    file_name="face_groups_report.xlsx",
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    use_container_width=True
                )
                
        except Exception as e:
            st.error(f"An error occurred: {str(e)}")
            progress_bar.empty()
            status_text.empty()

# Results display
if st.session_state.processed and st.session_state.grouped_faces:
    st.markdown("---")
    st.markdown("### 📋 Grouped Faces Preview")
    
    groups = st.session_state.grouped_faces
    
    # Create tabs for each group
    tabs = st.tabs([f"Group {i+1} ({len(group)} faces)" for i, group in enumerate(groups)])
    
    for idx, (tab, group) in enumerate(zip(tabs, groups)):
        with tab:
            cols = st.columns(3)
            for img_idx, image_name in enumerate(group):
                with cols[img_idx % 3]:
                    # Find the image in uploaded files
                    for file in uploaded_files:
                        if file.name == image_name:
                            image = Image.open(file)
                            st.image(image, caption=image_name, use_column_width=True)
                            file.seek(0)  # Reset file pointer
                            break

# Footer
st.markdown("---")
st.markdown("""
<div class="footer">
    <p>Designed by Bijay Paswan | v1.0 | Face Grouping Professional</p>
</div>
""", unsafe_allow_html=True)
